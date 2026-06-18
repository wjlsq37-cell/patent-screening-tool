from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill

from backend.core.config_manager import OUTPUT_DIR, ensure_data_dirs, load_config


RESULT_COLUMNS = [
    "排名",
    "综合评分",
    "相关度评分",
    "法律状态评分",
    "申请人评分",
    "时间评分",
    "推荐等级",
    "专利名称",
    "申请号",
    "公开号",
    "申请人",
    "发明人",
    "申请日",
    "公开日",
    "授权日",
    "专利类型",
    "法律状态",
    "剩余保护期估算",
    "IPC 分类号",
    "命中关键词",
    "AI 简述",
    "解决的问题",
    "实现方式",
    "核心发明点",
    "与用户需求的关系",
    "可能规避方向",
    "阅读建议",
    "详情链接",
    "人工备注",
]


def _sheet_name(name: str) -> str:
    cleaned = re.sub(r"[:\\/?*\[\]]", "_", name)
    return cleaned[:31] or "Sheet"


def _records_df(records: list[dict[str, Any]]) -> pd.DataFrame:
    normalized = []
    for record in records:
        item = {column: record.get(column, "") for column in RESULT_COLUMNS}
        normalized.append(item)
    return pd.DataFrame(normalized, columns=RESULT_COLUMNS)


def _keyword_df(requirement: str, keyword_analysis: dict[str, Any] | None, warnings: list[str] | None) -> pd.DataFrame:
    rows = [{"项目": "用户需求", "内容": requirement}]
    for key, value in (keyword_analysis or {}).items():
        if isinstance(value, list):
            value = "；".join(str(item) for item in value)
        rows.append({"项目": key, "内容": str(value)})
    if warnings:
        rows.append({"项目": "分析提示", "内容": "；".join(warnings)})
    return pd.DataFrame(rows, columns=["项目", "内容"])


def _format_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    score_columns = {"综合评分", "相关度评分", "法律状态评分", "申请人评分", "时间评分"}
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        header_lookup = {cell.value: cell.column for cell in ws[1]}
        for column_name in score_columns:
            col_idx = header_lookup.get(column_name)
            if col_idx:
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    row[0].number_format = "0.0"

        link_col = header_lookup.get("详情链接")
        if link_col:
            for row in ws.iter_rows(min_row=2, min_col=link_col, max_col=link_col):
                cell = row[0]
                if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"

        for column_cells in ws.columns:
            values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
            max_len = max((len(value) for value in values), default=10)
            letter = column_cells[0].column_letter
            ws.column_dimensions[letter].width = max(10, min(max_len + 2, 42))
    wb.save(path)


def export_analysis(
    records: list[dict[str, Any]],
    requirement: str,
    keyword_analysis: dict[str, Any] | None,
    warnings: list[str] | None = None,
) -> tuple[Path, str]:
    ensure_data_dirs()
    config = load_config()
    prefix = config.get("export", {}).get("filename_prefix", "PatentHub_AI分析")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{prefix}_{timestamp}.xlsx"
    path = OUTPUT_DIR / filename

    threshold = float(config.get("analysis", {}).get("high_relevance_threshold", 75))
    all_df = _records_df(records)
    sheets: dict[str, pd.DataFrame] = {
        "全部专利汇总": all_df,
        "高相关专利": _records_df([r for r in records if float(r.get("综合评分", 0) or 0) >= threshold]),
        "有效_授权专利": _records_df([r for r in records if r.get("法律状态分类") == "有效_授权专利"]),
        "审中_实质审查": _records_df([r for r in records if r.get("法律状态分类") == "审中_实质审查"]),
        "失效_终止_届满": _records_df([r for r in records if r.get("法律状态分类") == "失效_终止_届满"]),
        "驳回_撤回": _records_df([r for r in records if r.get("法律状态分类") == "驳回_撤回"]),
        "法律状态不明": _records_df([r for r in records if r.get("法律状态分类") == "法律状态不明"]),
        "重点申请人": _records_df([r for r in records if r.get("是否重点申请人")]),
    }

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet, df in sheets.items():
            if "综合评分" in df.columns:
                df = df.sort_values("综合评分", ascending=False)
            df.to_excel(writer, sheet_name=_sheet_name(sheet), index=False)
        _keyword_df(requirement, keyword_analysis, warnings).to_excel(
            writer, sheet_name=_sheet_name("检索关键词与分析说明"), index=False
        )
    _format_workbook(path)
    return path, filename

