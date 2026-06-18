from __future__ import annotations

from pathlib import Path
from typing import Any
import re

from openpyxl import load_workbook
import pandas as pd

from backend.core.field_mapper import STANDARD_FIELDS, auto_map_fields, validate_mapping


def _json_safe(value: Any) -> Any:
    if pd.isna(value):
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value) if not isinstance(value, (int, float, bool)) else value


def list_sheets(path: Path) -> list[str]:
    excel = pd.ExcelFile(path)
    return list(excel.sheet_names)


def _parse_hyperlink_formula(value: Any) -> tuple[str, str] | None:
    if not isinstance(value, str):
        return None
    text = value.strip()
    match = re.match(r'^=HYPERLINK\(\s*"((?:[^"]|"")*)"\s*[,;]\s*"((?:[^"]|"")*)"\s*\)\s*$', text, re.I)
    if not match:
        return None
    url = match.group(1).replace('""', '"')
    display = match.group(2).replace('""', '"')
    return display, url


def _apply_excel_links(path: Path, sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
    wb = load_workbook(path, read_only=False, data_only=False)
    ws = wb[sheet_name]
    df = df.astype(object)
    links: dict[tuple[Any, str], str] = {}

    for row_pos in range(len(df)):
        excel_row = row_pos + 2
        for col_pos, column in enumerate(df.columns):
            cell = ws.cell(row=excel_row, column=col_pos + 1)
            parsed_formula = _parse_hyperlink_formula(cell.value)
            display = ""
            url = ""

            if parsed_formula:
                display, url = parsed_formula
            elif cell.hyperlink:
                url = cell.hyperlink.target or cell.hyperlink.location or ""
                display = cell.hyperlink.display or str(cell.value or "")

            if display:
                df.iat[row_pos, col_pos] = display
            if url:
                links[(df.index[row_pos], str(column))] = url

    df.attrs["hyperlinks"] = links
    return df


def read_sheet(path: Path, sheet_name: str | None = None) -> pd.DataFrame:
    sheets = list_sheets(path)
    if not sheets:
        raise ValueError("xlsx 文件不包含任何 sheet。")
    active_sheet = sheet_name if sheet_name in sheets else sheets[0]
    df = pd.read_excel(path, sheet_name=active_sheet)
    df.columns = [str(col).strip() for col in df.columns]
    df = _apply_excel_links(path, active_sheet, df)
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
    if df.empty:
        raise ValueError("xlsx 文件为空，未读取到有效数据。")
    return df


def preview_dataframe(df: pd.DataFrame, limit: int = 10) -> dict[str, Any]:
    preview = df.head(limit).copy()
    rows = []
    for _, row in preview.iterrows():
        rows.append({str(col): _json_safe(row[col]) for col in preview.columns})
    columns = [str(col) for col in df.columns]
    return {
        "columns": columns,
        "rows": rows,
        "row_count": int(len(df)),
        "field_mapping": auto_map_fields(columns),
    }


def standardize_records(df: pd.DataFrame, mapping: dict[str, str]) -> list[dict[str, Any]]:
    mapping = validate_mapping(mapping, [str(col) for col in df.columns])
    links = df.attrs.get("hyperlinks", {})
    records: list[dict[str, Any]] = []
    for row_index, row in df.iterrows():
        item: dict[str, Any] = {}
        for field in STANDARD_FIELDS:
            source = mapping.get(field, "")
            item[field] = _json_safe(row[source]) if source else ""
        if not item.get("详情链接"):
            for linked_field in ("公开号", "申请号"):
                source = mapping.get(linked_field, "")
                link = links.get((row_index, source))
                if link:
                    item["详情链接"] = link
                    break
        records.append(item)
    return records
